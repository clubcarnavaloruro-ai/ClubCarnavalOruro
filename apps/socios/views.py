from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from .models import Socio
from .models import UserProfile
from .models import generar_codigo_socio
from django.contrib.auth import update_session_auth_hash
import csv
from io import TextIOWrapper
from django.contrib.auth.models import User
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def listar_socios(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '').strip()

    socios = Socio.objects.select_related('user')
    if q:
        socios = socios.filter(
            Q(nombre__icontains=q)
            | Q(apellido__icontains=q)
            | Q(apellido_paterno__icontains=q)
            | Q(apellido_materno__icontains=q)
            | Q(email__icontains=q)
            | Q(user__username__icontains=q)
            | Q(carnet_ci__icontains=q)
        )
    if estado:
        socios = socios.filter(estado=estado)

    paginator = Paginator(socios.order_by('-fecha_ingreso'), 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Asegurar que cada usuario listado tenga un UserProfile para evitar errores en plantillas
    for s in page_obj.object_list:
        try:
            UserProfile.objects.get_or_create(user=s.user)
        except Exception:
            pass

    return render(request, 'socios/socios.html', {
        'page_obj': page_obj,
        'q': q,
        'estado': estado,
        'is_admin': request.user.is_staff,
    })


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def crear_socio(request):
    if request.method != 'POST':
        return redirect('socios:listar_socios')

    username = request.POST.get('username', '').strip()
    nombre = request.POST.get('nombre', '').strip()
    apellido_paterno = request.POST.get('apellido_paterno', '').strip()
    apellido_materno = request.POST.get('apellido_materno', '').strip()
    apellido = request.POST.get('apellido', '').strip() or f"{apellido_paterno} {apellido_materno}".strip()
    email = request.POST.get('email', '').strip()
    telefono = request.POST.get('telefono', '').strip()
    ciudad = request.POST.get('ciudad', '').strip()
    direccion = request.POST.get('direccion', '').strip()
    password = request.POST.get('password', '')
    fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip() or None
    razon = request.POST.get('razon', '').strip()
    carnet_ci = request.POST.get('carnet_ci', '').strip()
    carnet_complemento = request.POST.get('carnet_complemento', '').strip()
    observacion = request.POST.get('observacion', '').strip()

    if not username or not nombre or not (apellido_paterno or apellido) or not email or not password:
        messages.error(request, 'Completa los campos obligatorios.')
        return redirect('socios:listar_socios')

    if User.objects.filter(username=username).exists():
        messages.error(request, 'El nombre de usuario ya existe.')
        return redirect('socios:listar_socios')

    user = User.objects.create_user(username=username, email=email, password=password)
    user.first_name = nombre
    user.last_name = apellido_paterno or apellido
    user.save()
    Socio.objects.create(
        user=user,
        codigo_socio=generar_codigo_socio(),
        nombre=nombre,
        apellido_paterno=apellido_paterno,
        apellido_materno=apellido_materno,
        apellido=apellido,
        email=email,
        telefono=telefono,
        ciudad=ciudad,
        direccion=direccion,
        fecha_nacimiento=fecha_nacimiento,
        razon=razon,
        carnet_ci=carnet_ci,
        carnet_complemento=carnet_complemento,
        observacion=observacion,
    )
    messages.success(request, 'Socio registrado correctamente.')
    return redirect('socios:listar_socios')


@login_required
def perfil_socio(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    socio = None
    try:
        socio = user.socio_profile
    except Socio.DoesNotExist:
        socio = None

    entregas = socio.entregas_souvenir.select_related('entregado_por').all() if socio else []
    paginator = Paginator(entregas, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'socios/perfil.html', {
        'socio': socio,
        'page_obj': page_obj,
        'is_admin': request.user.is_staff,
        'user_profile': profile,
    })


@login_required
def editar_perfil(request):
    if request.method != 'POST':
        return redirect('socios:perfil_socio')

    user = request.user
    email = request.POST.get('email', '').strip()
    telefono = request.POST.get('telefono', '').strip()

    # Actualizar email en User
    if email:
        user.email = email
        user.save()

    # Actualizar datos en Socio si existe
    try:
        socio = user.socio_profile
        socio.email = email or socio.email
        socio.telefono = telefono or socio.telefono
        socio.save()
    except Socio.DoesNotExist:
        pass

    messages.success(request, 'Datos de perfil actualizados.')
    return redirect('socios:perfil_socio')


@login_required
def cambiar_contrasena(request):
    if request.method != 'POST':
        return redirect('socios:perfil_socio')

    user = request.user
    current = request.POST.get('current_password', '')
    new1 = request.POST.get('new_password1', '')
    new2 = request.POST.get('new_password2', '')

    # Validar que los tres campos estén presentes
    if not current or not new1 or not new2:
        messages.error(request, 'Completa los 3 campos requeridos para cambiar la contraseña.')
        return redirect('socios:perfil_socio')

    if not user.check_password(current):
        messages.error(request, 'La contraseña actual es incorrecta.')
        return redirect('socios:perfil_socio')

    if new1 != new2:
        messages.error(request, 'Las nuevas contraseñas no coinciden.')
        return redirect('socios:perfil_socio')

    try:
        user.set_password(new1)
        user.save()
        # Mantener la sesión activa
        update_session_auth_hash(request, user)
        messages.success(request, 'Contraseña actualizada correctamente.')
    except Exception:
        messages.error(request, 'No se pudo actualizar la contraseña.')

    return redirect('socios:perfil_socio')


@login_required
def subir_foto(request):
    if request.method != 'POST':
        return redirect('socios:perfil_socio')

    foto = request.FILES.get('foto')
    user_id = request.POST.get('user_id')

    # Si el usuario es admin puede subir foto para otro usuario
    if user_id and request.user.is_staff:
        from django.contrib.auth.models import User
        target = User.objects.filter(id=user_id).first()
        if not target:
            messages.error(request, 'Usuario no encontrado.')
            return redirect('socios:perfil_socio')
        profile, _ = UserProfile.objects.get_or_create(user=target)
    else:
        profile, _ = UserProfile.objects.get_or_create(user=request.user)

    if foto:
        profile.foto = foto
        profile.save()
        messages.success(request, 'Foto de perfil actualizada.')
    else:
        messages.error(request, 'No se recibió archivo.')

    return redirect('socios:perfil_socio')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def crear_admin(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        if not username or not email or not password:
            messages.error(request, 'Completa los campos obligatorios.')
            return redirect('socios:crear_admin')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya existe.')
            return redirect('socios:crear_admin')
        user = User.objects.create_user(username=username, email=email, password=password)
        user.first_name = first_name
        user.last_name = last_name
        user.is_staff = True
        user.save()
        messages.success(request, 'Administrador creado correctamente.')
        return redirect('socios:listar_admins')
    return redirect('socios:listar_admins')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def importar_socios(request):
    if request.method == 'POST':
        f = request.FILES.get('file')
        if not f:
            messages.error(request, 'Sube un archivo CSV.')
            return redirect('socios:listar_socios')
        try:
            text = TextIOWrapper(f.file, encoding='utf-8')
            reader = csv.DictReader(text)
            created = 0
            for row in reader:
                username = row.get('username') or row.get('usuario') or ''
                nombre = row.get('nombre') or ''
                apellido_paterno = row.get('apellido_paterno') or row.get('apellido') or ''
                apellido_materno = row.get('apellido_materno') or ''
                apellido = row.get('apellido') or f"{apellido_paterno} {apellido_materno}".strip()
                email = row.get('email') or ''
                password = row.get('password') or User.objects.make_random_password()
                telefono = row.get('telefono') or ''
                ciudad = row.get('ciudad') or ''
                direccion = row.get('direccion') or ''
                fecha_nacimiento = row.get('fecha_nacimiento') or None
                razon = row.get('razon') or ''
                carnet_ci = row.get('carnet_ci') or ''
                carnet_complemento = row.get('carnet_complemento') or ''
                if not username or User.objects.filter(username=username).exists():
                    continue
                user = User.objects.create_user(username=username, email=email, password=password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                Socio.objects.create(user=user, codigo_socio=generar_codigo_socio(), nombre=nombre, apellido_paterno=apellido_paterno, apellido_materno=apellido_materno, apellido=apellido, email=email, telefono=telefono, ciudad=ciudad, direccion=direccion, fecha_nacimiento=fecha_nacimiento, razon=razon, carnet_ci=carnet_ci, carnet_complemento=carnet_complemento)
                created += 1
            messages.success(request, f'Socios importados: {created}')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {e}')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def importar_socios_masivo(request):
    return render(request, 'socios/importar_masivo.html')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def importar_socios_xlsx_preview(request):
    if request.method != 'POST':
        return redirect('socios:importar_socios_masivo')

    f = request.FILES.get('file')
    if not f:
        messages.error(request, 'Sube un archivo .xlsx.')
        return redirect('socios:importar_socios_masivo')

    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, 'El archivo está vacío.')
            return redirect('socios:importar_socios_masivo')

        headers = [str(cell or '') for cell in rows[0]]
        preview = [[str(cell or '') for cell in row] for row in rows[1:11]]
        
        def convert_cell_value(cell):
            if cell is None:
                return ''
            if isinstance(cell, datetime):
                return cell.strftime('%Y-%m-%d') if cell.time() == datetime.min.time() else cell.strftime('%Y-%m-%d %H:%M:%S')
            return str(cell)
        
        preview_data = [
            [convert_cell_value(cell) for cell in row[:14]]
            for row in rows[1:]
            if any(cell is not None for cell in row[:14])
        ]

        request.session['socios_import_preview'] = preview_data

        return render(request, 'socios/importar_masivo.html', {
            'preview_headers': headers,
            'preview_rows': preview,
        })
    except Exception as e:
        messages.error(request, f'Error al procesar xlsx: {e}')
        return redirect('socios:importar_socios_masivo')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def importar_socios_xlsx_confirm(request):
    preview_data = request.session.pop('socios_import_preview', None)
    if not preview_data:
        messages.error(request, 'No hay datos para confirmar.')
        return redirect('socios:importar_socios_masivo')

    created = 0
    skipped = 0
    errors = []
    for row in preview_data:
        vals = [(c or '') for c in row[:14]]
        username = vals[0]
        nombre = vals[1]
        apellido_paterno = vals[2]
        apellido_materno = vals[3]
        apellido = vals[4] or f"{apellido_paterno} {apellido_materno}".strip()
        email = vals[5]
        password = vals[6]
        telefono = vals[7]
        ciudad = vals[8]
        direccion = vals[9]
        fecha_nacimiento = vals[10] or None
        razon = vals[11]
        carnet_ci = vals[12]
        carnet_complemento = vals[13]
        
        if not username:
            skipped += 1
            errors.append(f"Fila sin username: {nombre}")
            continue
        
        if not password:
            password = User.objects.make_random_password()
        
        try:
            if User.objects.filter(username=username).exists():
                # Actualizar usuario existente
                user = User.objects.get(username=username)
                user.email = email
                if password and password != '':
                    user.set_password(password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                
                # Actualizar o crear socio
                socio, created_socio = Socio.objects.update_or_create(
                    user=user,
                    defaults={
                        'nombre': nombre,
                        'apellido_paterno': apellido_paterno,
                        'apellido_materno': apellido_materno,
                        'apellido': apellido,
                        'email': email,
                        'telefono': telefono,
                        'ciudad': ciudad,
                        'direccion': direccion,
                        'fecha_nacimiento': fecha_nacimiento,
                        'razon': razon,
                        'carnet_ci': carnet_ci,
                        'carnet_complemento': carnet_complemento,
                    }
                )
                if created_socio:
                    created += 1
                else:
                    created += 1  # Contar como actualizado
            else:
                # Crear nuevo usuario
                user = User.objects.create_user(username=username, email=email, password=password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                Socio.objects.create(user=user, codigo_socio=generar_codigo_socio(), nombre=nombre, apellido_paterno=apellido_paterno, apellido_materno=apellido_materno, apellido=apellido, email=email, telefono=telefono, ciudad=ciudad, direccion=direccion, fecha_nacimiento=fecha_nacimiento, razon=razon, carnet_ci=carnet_ci, carnet_complemento=carnet_complemento)
                created += 1
        except Exception as e:
            skipped += 1
            errors.append(f"Error creando {username}: {str(e)}")

    if errors:
        messages.warning(request, f'Socios importados: {created}, omitidos: {skipped}. Errores: {"; ".join(errors[:5])}')
    else:
        messages.success(request, f'Socios importados desde XLSX: {created}')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def descargar_plantilla_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'socios'
    
    # Encabezados
    headers = ['username', 'nombre', 'apellido_paterno', 'apellido_materno', 'apellido', 'email', 'password', 'telefono', 'ciudad', 'direccion', 'fecha_nacimiento', 'razon', 'carnet_ci', 'carnet_complemento']
    ws.append(headers)
    
    # Estilo para el encabezado
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Azul
    header_font = Font(color="FFFFFF", bold=True, size=11)  # Letra blanca y negrita
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 15,  # username
        'B': 20,  # nombre
        'C': 20,  # apellido_paterno
        'D': 20,  # apellido_materno
        'E': 25,  # apellido
        'F': 25,  # email
        'G': 15,  # password
        'H': 15,  # telefono
        'I': 15,  # ciudad
        'J': 30,  # direccion
        'K': 15,  # fecha_nacimiento
        'L': 30,  # razon
        'M': 15,  # carnet_ci
        'N': 15,  # carnet_complemento
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Añadir bordes a todas las celdas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Ejemplo de fila con estilo
    example_row = ['jdoe', 'Juan', 'Perez', 'Gomez', 'Perez Gomez', 'jdoe@example.com', 'Passw0rd!', '71234567', 'Oruro', 'Dirección 123', '1990-01-01', 'Quiero participar', '1234567', '-1A']
    ws.append(example_row)
    
    # Aplicar bordes y colores alternados a las filas de datos
    for row_num in range(2, ws.max_row + 1):
        # Color alternado para filas
        if row_num % 2 == 0:
            row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Azul claro
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Congelar la primera fila (encabezado)
    ws.freeze_panes = "A2"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=socios_plantilla.xlsx'
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def importar_socios_xlsx(request):
    if request.method == 'POST':
        f = request.FILES.get('file')
        if not f:
            messages.error(request, 'Sube un archivo .xlsx')
            return redirect('socios:listar_socios')
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            created = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                vals = [ (c or '') for c in row[:14] ]
                username = vals[0]
                nombre = vals[1]
                apellido_paterno = vals[2]
                apellido_materno = vals[3]
                apellido = vals[4] or f"{apellido_paterno} {apellido_materno}".strip()
                email = vals[5]
                password = vals[6]
                telefono = vals[7]
                ciudad = vals[8]
                direccion = vals[9]
                fecha_nacimiento = vals[10] or None
                razon = vals[11]
                carnet_ci = vals[12]
                carnet_complemento = vals[13]
                if not username or User.objects.filter(username=username).exists():
                    continue
                if not password:
                    password = User.objects.make_random_password()
                user = User.objects.create_user(username=username, email=email, password=password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                Socio.objects.create(user=user, codigo_socio=generar_codigo_socio(), nombre=nombre, apellido_paterno=apellido_paterno, apellido_materno=apellido_materno, apellido=apellido, email=email, telefono=telefono, ciudad=ciudad, direccion=direccion, fecha_nacimiento=fecha_nacimiento, razon=razon, carnet_ci=carnet_ci, carnet_complemento=carnet_complemento)
                created += 1
            messages.success(request, f'Socios importados desde XLSX: {created}')
        except Exception as e:
            messages.error(request, f'Error al procesar xlsx: {e}')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def listar_admins(request):
    q = request.GET.get('q', '').strip()
    activo = request.GET.get('activo', '').strip()
    admins = User.objects.filter(is_staff=True)

    if q:
        admins = admins.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
        )

    if activo == 'si':
        admins = admins.filter(is_active=True)
    elif activo == 'no':
        admins = admins.filter(is_active=False)

    paginator = Paginator(admins.order_by('username'), 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'admins/admins.html', {'page_obj': page_obj, 'q': q, 'activo': activo})


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def ver_admin(request, user_id):
    return redirect('socios:listar_admins')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def editar_admin(request, user_id):
    user = get_object_or_404(User, id=user_id, is_staff=True)
    if request.method == 'POST':
        user.username = request.POST.get('username', user.username).strip()
        user.email = request.POST.get('email', user.email).strip()
        user.first_name = request.POST.get('first_name', user.first_name).strip()
        user.last_name = request.POST.get('last_name', user.last_name).strip()
        password = request.POST.get('password')
        if password:
            user.set_password(password)
        user.save()
        messages.success(request, 'Administrador actualizado.')
        return redirect('socios:listar_admins')
    return redirect('socios:listar_admins')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def eliminar_admin(request, user_id):
    user = get_object_or_404(User, id=user_id, is_staff=True)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Administrador eliminado.')
        return redirect('socios:listar_admins')
    return redirect('socios:listar_admins')


@login_required
def mis_souvenirs(request):
    try:
        socio = request.user.socio_profile
    except Socio.DoesNotExist:
        messages.error(request, 'No se encontró perfil de socio.')
        return redirect('/')
    entregas = socio.entregas_souvenir.select_related('souvenir', 'entregado_por').all()
    paginator = Paginator(entregas, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'socios/mis_souvenirs.html', {'page_obj': page_obj})


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def editar_socio(request, socio_id):
    if request.method != 'POST':
        return redirect('socios:listar_socios')

    socio = get_object_or_404(Socio, id=socio_id)
    socio.nombre = request.POST.get('nombre', '').strip()
    socio.apellido_paterno = request.POST.get('apellido_paterno', '').strip()
    socio.apellido_materno = request.POST.get('apellido_materno', '').strip()
    socio.apellido = request.POST.get('apellido', '').strip() or f"{socio.apellido_paterno} {socio.apellido_materno}".strip()
    socio.email = request.POST.get('email', '').strip()
    socio.telefono = request.POST.get('telefono', '').strip()
    socio.ciudad = request.POST.get('ciudad', '').strip()
    socio.direccion = request.POST.get('direccion', '').strip()
    socio.carnet_ci = request.POST.get('carnet_ci', '').strip()
    socio.carnet_complemento = request.POST.get('carnet_complemento', '').strip()
    socio.observacion = request.POST.get('observacion', '').strip()
    socio.save()
    messages.success(request, 'Datos del socio actualizados.')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def activar_socio(request, socio_id):
    socio = get_object_or_404(Socio, id=socio_id)
    socio.estado = 'activo'
    socio.save()
    messages.success(request, 'Socio activado.')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def desactivar_socio(request, socio_id):
    socio = get_object_or_404(Socio, id=socio_id)
    socio.estado = 'inactivo'
    socio.save()
    messages.success(request, 'Socio desactivado.')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def historial_souvenirs(request, socio_id):
    socio = get_object_or_404(Socio, id=socio_id)
    entregas = socio.entregas_souvenir.select_related('souvenir', 'evento', 'entregado_por').order_by('-fecha_entrega')
    total_entregas = entregas.count()
    paginator = Paginator(entregas, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'socios/historial_souvenirs.html', {
        'socio': socio,
        'page_obj': page_obj,
        'total_entregas': total_entregas,
    })


@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/login/')
def eliminar_socio(request, socio_id):
    socio = get_object_or_404(Socio, id=socio_id)
    socio.user.delete()
    socio.delete()
    messages.success(request, 'Socio eliminado definitivamente.')
    return redirect('socios:listar_socios')
